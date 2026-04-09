#!/usr/bin/env python3
"""xlsx_sync.py — 구글드라이브 자산계산기.xlsx → portfolio.json 자동 동기화
                 + 실시간 시세 자동 업데이트 (G열)

단독 실행:  python3 xlsx_sync.py
다른 스크립트에서 호출: from xlsx_sync import load_portfolio
"""

import os, json, re, warnings, threading, platform
warnings.filterwarnings('ignore')

_sys = platform.system()
if _sys == "Windows":
    XLSX_PATH = r"G:\내 드라이브\PF\자산 계산기(클로드).xlsx"
elif _sys == "Darwin":  # macOS
    XLSX_PATH = os.path.expanduser(
        "~/Library/CloudStorage/GoogleDrive-miyoo1016@gmail.com"
        "/내 드라이브/PF/자산 계산기(클로드).xlsx"
    )
else:  # Android(Termux) / Linux — Google Drive 없음, portfolio.json 직접 사용
    XLSX_PATH = ""
PORTFOLIO_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "portfolio.json")
SHEET_NAME = "📊 자산 계산기"

# 종목명 → Yahoo Finance 티커 매핑
# GOLD_KRX  : 네이버 금융 KRX 금현물 시세 직접 조회 (GC=F 계산 아님)
# XLSX_PRICE : G열 현재가를 KS 티커로 자동 업데이트
TICKER_MAP = {
    # ── 미국 주식 (yfinance 실시간) ───────────────────────────
    "QQQM":                        "QQQM",
    "Alphabet A":                  "GOOGL",
    "Google":                      "GOOGL",
    "GOOGL":                       "GOOGL",
    "SPY":                         "SPY",
    "GLD":                         "GLD",
    "BTC":                         "BTC-USD",
    "QQQ":                         "QQQ",
    # ── 국내 주식 / ETF (yfinance KS 티커) ───────────────────
    "삼성전자":                      "005930.KS",
    "TIGER CD금리(합성)":           "357870.KS",   # TIGER CD금리투자KIS(합성)
    "TIGER CD금리투자KIS(합성)":    "357870.KS",
    # ── 국내 ETF (KS 티커 — 시세 자동 업데이트) ──────────────
    "KODEX 나스닥100":              "XLSX_PRICE",  # → 379810.KS
    "KODEX 미국나스닥100":          "XLSX_PRICE",  # → 379810.KS
    "KODEX S&P500":                "XLSX_PRICE",  # → 379800.KS
    "KODEX 미국S&P500":            "XLSX_PRICE",  # → 379800.KS
    "KODEX 미국반도체":             "XLSX_PRICE",  # → 390390.KS
    # ── 원자재 (KRX 금현물 — 네이버 금융) ────────────────────
    "금현물(KRX)":                  "GOLD_KRX",
    # ── 현금 ─────────────────────────────────────────────────
    "현금":                        "CASH",
}

# 자산구분 → 통화 매핑
CURRENCY_MAP = {
    "미국주식(달러)":  "USD",
    "국내일반(원화)":  "KRW",
    "국내ETF(원화)":   "KRW",
    "원자재(원화)":    "KRW",
    "현금성(원화)":    "KRW",
    "현금성(달러)":    "USD",
    "은행현금(원화)":  "KRW",
    "은행현금(달러)":  "USD",
    "CMA(원화)":       "KRW",
}

# 현금 자산구분 목록
CASH_TYPES = {"은행현금(원화)", "은행현금(달러)", "현금성(원화)", "현금성(달러)", "CMA(원화)"}

# XLSX_PRICE 종목 → 실제 yfinance KS 티커 (시세 쓰기용)
# 코드 출처: 사용자 직접 확인 + 네이버금융 자동완성 API
KS_TICKER_MAP = {
    "KODEX 나스닥100":     "379810.KS",   # KODEX 미국나스닥100
    "KODEX 미국나스닥100": "379810.KS",
    "KODEX S&P500":       "379800.KS",   # KODEX 미국S&P500
    "KODEX 미국S&P500":   "379800.KS",
    "KODEX 미국반도체":    "390390.KS",   # KODEX 미국반도체
}


def read_xlsx():
    """xlsx에서 보유 종목 + 현금 데이터 추출"""
    if not XLSX_PATH:  # Android/Termux
        return None

    try:
        import pandas as pd
    except ImportError:
        print("  pandas 필요: pip3 install pandas openpyxl")
        return None

    if not os.path.exists(XLSX_PATH):
        print(f"  파일 없음: {XLSX_PATH}")
        print("  구글드라이브가 연결되어 있는지 확인하세요.")
        return None

    try:
        df = pd.read_excel(XLSX_PATH, sheet_name=SHEET_NAME, header=None, engine='openpyxl')
    except Exception as e:
        print(f"  xlsx 읽기 실패: {e}")
        return None

    holdings = []
    # ── [NEW] L열 헤더 확인 및 강제 너비/스타일 조정 (9행) ──────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME]
        
        # 잘못 들어간 H1 삭제
        if ws.cell(row=1, column=8).value == "매수원가(₩)":
            ws.cell(row=1, column=8).value = None
        
        # L열 9행(Table Header) 제목 및 스타일 강제 적용
        l_header = ws.cell(row=9, column=12)
        l_header.value = "매입환율"
        l_header.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        l_header.fill = openpyxl.styles.PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        l_header.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # 너비 강제 조정 (20으로 넉넉하게)
        ws.column_dimensions['L'].width = 20
        
        # ── [NEW] 아래 데이터 칸들 스타일 복사 (K열 -> L열) ──────────
        # 10행부터 데이터가 있는 구석구석까지 스타일 복제
        for r_idx in range(10, 101): # 100행까지 넉넉하게 적용
            source_cell = ws.cell(row=r_idx, column=11) # K열(11)
            target_cell = ws.cell(row=r_idx, column=12) # L열(12)
            
            # 스타일 복사 (배경색, 테두리 등)
            if source_cell.has_style:
                from copy import copy
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
        
        wb.save(XLSX_PATH)
        print("  ✅ 엑셀 L열(매입환율)의 스타일이 기존 표와 동일하게 맞춰졌습니다.")
    except Exception as e:
        print(f"  ⚠️ 엑셀 스타일 조정 중 오류: {e}")

    # ── [NEW] 기준 환율(Base FX) 읽기: 14행 O열 (index 13, 14) ────────
    try:
        base_usdkrw = float(df.iloc[13, 14])
        if base_usdkrw != base_usdkrw or base_usdkrw <= 0:
            base_usdkrw = 1350.0  # fallback
    except Exception:
        base_usdkrw = 1350.0

    for _, row in df.iterrows():
        asset_type = str(row[1]).strip() if pd.notna(row[1]) else ""
        name       = str(row[2]).strip() if pd.notna(row[2]) else ""
        account    = str(row[3]).strip() if pd.notna(row[3]) else ""

        if not (asset_type and account):
            continue

        currency = CURRENCY_MAP.get(asset_type, "KRW")

        # ── 현금 처리 ────────────────────────────────────────────
        if asset_type in CASH_TYPES:
            # G열(index 6) = 현금 금액 직접 입력
            cash_val = row[6]
            try:
                cash_amt = float(cash_val)
            except (TypeError, ValueError):
                continue
            if cash_amt != cash_amt or cash_amt <= 0:   # NaN 또는 0 제외
                continue

            holdings.append({
                "name":       name if name else "현금",
                "ticker":     "CASH",
                "account":    account,
                "qty":        1,
                "avg_price":  round(cash_amt, 2),   # qty=1, avg=금액 → 평가금액=금액
                "currency":   currency,
                "asset_type": asset_type,
                "is_cash":    True,
            })
            continue

        # ── 일반 종목 처리 ────────────────────────────────────────
        if not name:
            continue
        qty_raw  = row[4]
        avg_raw  = row[5]
        xlsx_price_raw = row[6]   # G열 = 현재가 (VLOOKUP 자동값)
        try:
            # [NEW] L열(index 11)에서 정밀 매입환율(또는 원화 원가) 추출
            cost_krw_raw = row[11] if len(row) > 11 else None
            if pd.notna(cost_krw_raw):
                val_str = str(cost_krw_raw).replace(',', '').replace('원', '').strip()
                # 숫자와 점(.)만 남기기
                import re
                val_str = re.sub(r'[^0-9\.]', '', val_str)
                precision_cost_krw = float(val_str) if val_str else None
            else:
                precision_cost_krw = None
        except (TypeError, ValueError):
            precision_cost_krw = None

        try:
            qty = float(qty_raw)
            avg = float(avg_raw)
        except (TypeError, ValueError):
            continue
        if qty != qty or qty <= 0:
            continue
        if avg != avg or avg <= 0:
            continue

        # G열 현재가 (XLSX_PRICE 티커용 fallback)
        try:
            xlsx_price = float(xlsx_price_raw)
            if xlsx_price != xlsx_price or xlsx_price <= 0:
                xlsx_price = None
        except (TypeError, ValueError):
            xlsx_price = None

        ticker = TICKER_MAP.get(name, "")
        # XLSX_PRICE 플레이스홀더 → 실제 yfinance KS 티커로 교체
        if ticker == "XLSX_PRICE":
            ticker = KS_TICKER_MAP.get(name, "XLSX_PRICE")

        holdings.append({
            "name":        name,
            "ticker":      ticker,
            "account":     account,
            "qty":         qty,
            "avg_price":   round(avg, 4),
            "xlsx_price":  xlsx_price,   # xlsx의 최신 현재가 (fallback용)
            "currency":    currency,
            "asset_type":  asset_type,
            "is_cash":     False,
            "base_usdkrw": base_usdkrw,  # 기준 환율 추가
            "precision_cost_krw": precision_cost_krw, # [NEW] 정밀 원화 매수 원가
        })

    return holdings


def sync_to_json(holdings):
    """holdings 리스트 → portfolio.json 저장 (계좌별 그룹)"""
    accounts = {}
    for h in holdings:
        acc = h["account"]
        accounts.setdefault(acc, []).append({
            "name":       h["name"],
            "ticker":     h["ticker"],
            "qty":        h["qty"],
            "avg_price":  h["avg_price"],
            "currency":   h["currency"],
            "asset_type": h["asset_type"],
            "xlsx_price": h.get("xlsx_price"),
            "is_cash":    h.get("is_cash", False),
            "base_usdkrw": h.get("base_usdkrw"),
            "precision_cost_krw": h.get("precision_cost_krw"),
        })

    with open(PORTFOLIO_JSON, "w", encoding="utf-8") as f:
        json.dump(accounts, f, ensure_ascii=False, indent=2)

    return accounts

def safe_save_xlsx(wb, path):
    """임시 파일을 거쳐 안전하게 저장 (파일 손상 방지)"""
    import os
    tmp_path = path + ".tmp"
    try:
        wb.save(tmp_path)
        if os.path.exists(tmp_path):
            # 기존 파일 제거 후 교체
            if os.path.exists(path):
                os.remove(path)
            os.rename(tmp_path, path)
            return True
    except Exception as e:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        print(f"  ⚠️ 엑셀 저장 실패 (보안/접근권한 확인): {e}")
        return False
    return False


def update_xlsx_live_fx(usdkrw):
    """엑셀 O14 셀에 실시간 환율을 자동으로 기입"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME]
        
        # O14 셀 (row=14, column=15)
        fx_cell = ws.cell(row=14, column=15)
        fx_cell.value = usdkrw
        # 사용자 서식 유지 혹은 숫자형으로 강제
        fx_cell.number_format = '#,##0.00"원/달러"'
        
        if safe_save_xlsx(wb, XLSX_PATH):
            print(f"  ✅ 엑셀 O14 셀에 실시간 환율(₩{usdkrw:,.2f})이 업데이트되었습니다.")
    except Exception as e:
        print(f"  ⚠️ 엑셀 환율 업데이트 실패: {e}")


def load_portfolio():
    """
    다른 스크립트에서 호출.
    xlsx → holdings 리스트 반환. 실패 시 portfolio.json fallback.
    """
    holdings = read_xlsx()
    if holdings:
        sync_to_json(holdings)
        return holdings

    if os.path.exists(PORTFOLIO_JSON):
        with open(PORTFOLIO_JSON, encoding="utf-8") as f:
            data = json.load(f)
        flat = []
        for acc, items in data.items():
            for item in items:
                item["account"] = acc
                flat.append(item)
        return flat

    return []


def restore_vlookup_formulas():
    """G열 VLOOKUP 공식 복원 — xlsx_sync가 G열을 숫자로 덮어쓴 경우 1회 실행"""
    try:
        import openpyxl
    except ImportError:
        return False

    if not os.path.exists(XLSX_PATH):
        return False

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    restored = 0
    for row in ws.iter_rows():
        if len(row) < 7:
            continue
        asset_type = str(row[1].value or "").strip()
        name       = str(row[2].value or "").strip()
        if not asset_type or not name:
            continue
        if asset_type in CASH_TYPES:
            continue   # 현금행은 G열 직접 입력 — 건드리지 않음

        try:
            qty = float(row[4].value or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue

        # 이미 공식이면 스킵, 숫자(이전에 덮어쓴 것)면 복원
        g_cell = row[6]
        if isinstance(g_cell.value, str) and g_cell.value.startswith("="):
            continue   # 이미 공식 있음

        g_cell.value = f"=VLOOKUP(C{g_cell.row},$N$5:$O$12,2,0)"
        restored += 1

    if restored:
        if safe_save_xlsx(wb, XLSX_PATH):
            print(f"  🔧 G열 VLOOKUP 공식 {restored}개 복원 완료")
    return True


def fetch_and_write_prices():
    """실시간 시세를 조회해 xlsx N5:O12 룩업 테이블(O열)에 자동 업데이트
    ※ G열에는 VLOOKUP 공식이 있어야 함 — G열을 직접 건드리지 않음
    """
    try:
        import yfinance as yf
        import openpyxl
    except ImportError as e:
        print(f"  ⚠ 필요 패키지 없음: {e}  →  pip3 install yfinance openpyxl")
        return False

    if not os.path.exists(XLSX_PATH):
        print("  ⚠ 파일 없음 — 구글드라이브 연결 확인")
        return False

    # ── 1. 환율 조회 ────────────────────────────────────────────
    print("  환율·시세 조회 중...", end=" ", flush=True)
    try:
        tk_fx = yf.Ticker("USDKRW=X")
        usdkrw = tk_fx.fast_info.get('last_price') or tk_fx.fast_info.get('lastPrice')
        if not usdkrw:
            usdkrw = float(tk_fx.history(period="2d")["Close"].iloc[-1])
    except Exception:
        usdkrw = 1450.0
    print(f"USDKRW={usdkrw:,.1f}")

    # ── 2. 워크북 로드 + N5:O12 룩업 테이블 셀 매핑 ─────────────
    wb  = openpyxl.load_workbook(XLSX_PATH)
    ws  = wb[SHEET_NAME]

    # N열(col 14) = 종목명, O열(col 15) = 현재가 입력 셀  (행 5~12)
    n_to_o = {}
    for r in range(5, 13):
        n_val = str(ws.cell(row=r, column=14).value or "").strip()
        if n_val:
            n_to_o[n_val] = ws.cell(row=r, column=15)

    # ── 3. 조회할 티커 목록 수집 ─────────────────────────────────
    # (O열_셀, 종목명, 야후티커, 통화)
    targets = []
    seen_names = set()
    for row in ws.iter_rows():
        if len(row) < 7:
            continue
        asset_type = str(row[1].value or "").strip()
        name       = str(row[2].value or "").strip()
        if not asset_type or not name:
            continue
        if asset_type in CASH_TYPES:
            continue
        if name in seen_names:
            continue   # 동일 종목 중복 스킵

        try:
            qty = float(row[4].value or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue

        # 티커 결정
        ticker = TICKER_MAP.get(name, "")
        if ticker in ("", "XLSX_PRICE"):
            ticker = KS_TICKER_MAP.get(name, "")
        if not ticker or ticker == "CASH":
            continue

        # N5:O12 테이블에서 해당 종목 O열 셀 찾기
        o_cell = n_to_o.get(name)
        if o_cell is None:
            continue   # 룩업 테이블에 없으면 스킵

        currency = CURRENCY_MAP.get(asset_type, "KRW")
        targets.append((o_cell, name, ticker, currency))
        seen_names.add(name)

    if not targets:
        print("  업데이트할 종목 없음")
        return False

    # ── 4. 병렬 시세 조회 ───────────────────────────────────────
    regular_tickers = list({t for _, _, t, _ in targets if t != "GOLD_KRX"})
    price_data = {}
    errors     = []

    def _fetch_batch():
        try:
            if not regular_tickers:
                return
            data = yf.download(regular_tickers, period="2d",
                                auto_adjust=True, progress=False, threads=True)
            closes = data["Close"] if "Close" in data else data
            for tk in regular_tickers:
                try:
                    col = closes[tk] if tk in closes.columns else closes
                    val = float(col.dropna().iloc[-1])
                    price_data[tk] = val
                except Exception:
                    pass
        except Exception as e:
            errors.append(f"batch: {e}")

    def _fetch_gold():
        """KRX 금현물 — 네이버 모바일 증권 API (M04020000 = 국내 금, 한국거래소)"""
        try:
            import subprocess
            r = subprocess.run(
                ['curl', '-s', '-A', 'Mozilla/5.0',
                 'https://api.stock.naver.com/marketindex/metals/M04020000'],
                capture_output=True, timeout=10
            )
            d = json.loads(r.stdout.decode('utf-8', errors='replace'))
            price_str = d.get('closePrice') or d.get('currentPrice') or ''
            price = float(price_str.replace(',', ''))
            if price > 0:
                price_data["GOLD_KRX"] = int(round(price))
                return
            errors.append("금현물: 가격 없음")
        except Exception as e:
            errors.append(f"금현물: {e}")

    t1 = threading.Thread(target=_fetch_batch)
    t2 = threading.Thread(target=_fetch_gold)
    t1.start(); t2.start()
    t1.join();  t2.join()

    # ── 5. O열(룩업 테이블)에 시세 기록 + O14 환율 업데이트 ──────
    skipped_names = []
    written_names = set()
    print(f"\n  {'종목':<22} {'티커':<14} {'현재가':>14}")
    print("  " + "─" * 54)

    for o_cell, name, ticker, currency in targets:
        price = price_data.get(ticker)
        if price is None or price <= 0:
            skipped_names.append(name)
            continue

        # 원화 종목은 정수, USD는 소수 4자리
        if currency == "KRW":
            o_cell.value = int(round(price))
            disp = f"₩{int(round(price)):>12,}"
        else:
            o_cell.value = round(price, 4)
            disp = f"${price:>12.4f}"

        written_names.add(name)
        print(f"  {name:<22} {ticker:<14} {disp}")

    # 환율 O14 업데이트
    ws["O14"] = round(usdkrw, 2)

    # ── 6. 저장 ─────────────────────────────────────────────────
    if safe_save_xlsx(wb, XLSX_PATH):
        print("  " + "─" * 54)
        print(f"  ✅ {len(written_names)}개 종목 시세 + 환율({usdkrw:,.1f}원) 업데이트 완료")
    if skipped_names:
        uniq = list(dict.fromkeys(skipped_names))
        print(f"  ⚠ 조회 실패: {', '.join(uniq)}")
    if errors:
        print(f"  ⚠ 오류: {'; '.join(errors)}")
    return True


def main():
    print("\n" + "━"*56)
    print("  📥  xlsx → portfolio.json 동기화")
    print("━"*56)
    holdings = read_xlsx()
    if not holdings:
        print("  동기화 실패")
        return

    accounts = sync_to_json(holdings)

    stocks = [h for h in holdings if not h.get("is_cash")]
    cashes = [h for h in holdings if h.get("is_cash")]

    print(f"  완료: 종목 {len(stocks)}개 + 현금 {len(cashes)}개 = 총 {len(holdings)}개  ({len(accounts)}개 계좌)\n")
    print(f"  {'계좌':<22} {'종목':<18} {'구분':<10} {'금액/평단가':>14} {'통화'}")
    print("  " + "─" * 72)

    unmapped = []
    for acc, items in accounts.items():
        for it in items:
            if it.get("is_cash"):
                amt = f"₩{it['avg_price']:>13,.0f}" if it['currency'] == 'KRW' else f"${it['avg_price']:>13,.2f}"
                print(f"  {acc:<22} {it['name']:<18} {'현금':^10} {amt} {it['currency']}")
            else:
                if not it['ticker']:
                    unmapped.append(it['name'])
                ticker_disp = it['ticker'] if it['ticker'] else "⚠매핑없음"
                print(f"  {acc:<22} {it['name']:<18} {ticker_disp:<10} {it['avg_price']:>14,.2f} {it['currency']}")

    if unmapped:
        print(f"\n  ⚠ 티커 미매핑: {', '.join(set(unmapped))}")
        print("  → xlsx_sync.py의 TICKER_MAP에 추가하세요")

    print(f"\n  저장 완료: {PORTFOLIO_JSON}")

    # ── G열 VLOOKUP 공식 복원 (한 번만 필요, 이미 공식이면 스킵) ──
    print("\n" + "━"*56)
    print("  🔧  G열 VLOOKUP 공식 확인·복원")
    print("━"*56)
    restore_vlookup_formulas()

    # ── 실시간 시세 자동 업데이트 (O열 룩업 테이블) ──────────────
    print("\n" + "━"*56)
    print("  📡  실시간 시세 → xlsx O열(룩업 테이블) 자동 업데이트")
    print("━"*56)
    fetch_and_write_prices()
    print()


if __name__ == "__main__":
    main()
